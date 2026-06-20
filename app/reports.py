import io
import json
import os
from datetime import datetime

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.config import BASE_DIR
from app.database import get_history, get_statistics
from app.models import get_model_name, load_confusion_matrix, load_experiments


def _register_cyrillic_font():
    """
    Регистрирует шрифт с поддержкой кириллицы.
    Возвращает имя зарегистрированного шрифта для reportlab.
    """
    candidates = [
        # Проектные файлы (если пользователь добавит свой шрифт)
        os.path.join(BASE_DIR, "static", "fonts", "DejaVuSans.ttf"),
        os.path.join(BASE_DIR, "static", "fonts", "Arial.ttf"),
        # Linux
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        # macOS
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/Library/Fonts/Arial.ttf",
        # Windows
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\DejaVuSans.ttf",
    ]
    for path in candidates:
        if os.path.isfile(path):
            font_name = "FoodVisionUnicode"
            if font_name not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont(font_name, path))
            return font_name
    # Последний fallback: встроенный Helvetica (русский может отображаться некорректно)
    return "Helvetica"


def generate_excel():
    experiments = load_experiments()
    history = get_history(500)
    stats = get_statistics()

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Сравнение архитектур"
    ws1.append(
        [
            "Модель",
            "Accuracy",
            "Precision",
            "Recall",
            "F1",
            "Инференс (мс)",
            "Размер (МБ)",
        ]
    )
    for arch in experiments["architectures"]:
        ws1.append(
            [
                arch["name"],
                arch["accuracy"],
                arch["precision"],
                arch["recall"],
                arch["f1"],
                arch["inference_ms"],
                arch["model_size_mb"],
            ]
        )

    ws2 = wb.create_sheet("История запусков")
    ws2.append(["ID", "Файл", "Модель", "Класс", "Уверенность", "Время (мс)", "Дата"])
    for row in history:
        ws2.append(
            [
                row["id"],
                row["filename"],
                get_model_name(row["model_id"]),
                row["top1_class"],
                row["top1_confidence"],
                row["inference_ms"],
                row["created_at"],
            ]
        )

    ws3 = wb.create_sheet("Статистика")
    ws3.append(["Показатель", "Значение"])
    ws3.append(["Всего распознаваний", stats["total"]])
    ws3.append(["Средняя уверенность", stats["avg_confidence"]])
    ws3.append(["Среднее время (мс)", stats["avg_inference_ms"]])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_pdf():
    experiments = load_experiments()
    confusion = load_confusion_matrix()
    stats = get_statistics()
    history = get_history(10)
    font_name = _register_cyrillic_font()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2 * cm, leftMargin=2 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleRu", parent=styles["Heading1"], fontName=font_name, fontSize=16, spaceAfter=12
    )
    h2 = ParagraphStyle("HeadingRu", parent=styles["Heading2"], fontName=font_name)
    body = ParagraphStyle("BodyRu", parent=styles["BodyText"], fontName=font_name)

    story = []
    story.append(Paragraph("Отчёт по практике: распознавание блюд", title_style))
    story.append(
        Paragraph(
            f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Вариант 4",
            body,
        )
    )
    story.append(Spacer(1, 12))

    ds = experiments["dataset"]
    story.append(Paragraph("1. Данные", h2))
    story.append(
        Paragraph(
            f"Датасет: {ds['name']}. Классов: {ds['num_classes']}. "
            f"Train/Val/Test: {ds['train']}/{ds['validation']}/{ds['test']}.",
            body,
        )
    )
    story.append(Spacer(1, 8))

    story.append(Paragraph("2. Сравнение архитектур", h2))
    table_data = [
        ["Модель", "Acc", "F1", "мс", "МБ"],
    ]
    for arch in experiments["architectures"]:
        mark = " *" if arch["id"] == experiments["best_model"] else ""
        table_data.append(
            [
                arch["name"] + mark,
                f"{arch['accuracy']:.3f}",
                f"{arch['f1']:.3f}",
                str(arch["inference_ms"]),
                str(arch["model_size_mb"]),
            ]
        )
    t = Table(table_data, colWidths=[5 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4a6741")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
            ]
        )
    )
    story.append(t)
    story.append(Paragraph("* — лучшая модель (EfficientNet-B0)", body))
    story.append(Spacer(1, 8))

    story.append(Paragraph("3. Статистика демо-модуля", h2))
    story.append(Paragraph(f"Распознаваний: {stats['total']}", body))
    story.append(Paragraph(f"Средняя уверенность: {stats['avg_confidence']}", body))
    story.append(Spacer(1, 8))

    story.append(Paragraph("4. Последние запросы", h2))
    for item in history[:5]:
        story.append(
            Paragraph(
                f"• {item['top1_class']} ({item['top1_confidence']:.0%}) — {item['created_at']}",
                body,
            )
        )

    story.append(Spacer(1, 8))
    story.append(Paragraph("5. Анализ ошибок", h2))
    for err in experiments["error_analysis"]:
        story.append(
            Paragraph(
                f"• {err['pair']}: {err['description']} (ошибка {err['error_rate']:.0%})",
                body,
            )
        )

    doc.build(story)
    buffer.seek(0)
    return buffer


def save_history_json():
    from app.database import export_history_json

    path = os.path.join(BASE_DIR, "data", "history_export.json")
    with open(path, "w", encoding="utf-8") as f:
        f.write(export_history_json())
    return path
